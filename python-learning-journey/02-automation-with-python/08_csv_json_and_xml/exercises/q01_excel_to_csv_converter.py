"""
Question:
Write a program that takes an Excel workbook with multiple sheets and
creates a separate .csv file for each sheet, named after the sheet.

"""

import csv
import os
import openpyxl


def create_sample_workbook(filename="multi_sheet_data.xlsx"):
    wb = openpyxl.Workbook()

    sheet1 = wb.active
    sheet1.title = "Products"
    sheet1.append(["Product", "Price"])
    sheet1.append(["Notebook", 50])
    sheet1.append(["Pen", 10])

    sheet2 = wb.create_sheet("Customers")
    sheet2.append(["Name", "City"])
    sheet2.append(["Rahul", "Panipat"])
    sheet2.append(["Priya", "Delhi"])

    wb.save(filename)
    print(f"Created sample workbook '{filename}' with sheets: Products, Customers")


def excel_to_csv(filename, output_folder="csv_output"):
    os.makedirs(output_folder, exist_ok=True)
    wb = openpyxl.load_workbook(filename)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        csv_filename = os.path.join(output_folder, f"{sheet_name}.csv")

        with open(csv_filename, "w", newline="") as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Converted sheet '{sheet_name}' -> '{csv_filename}'")

    print(f"\nDone. {len(wb.sheetnames)} sheet(s) converted.")


if __name__ == "__main__":
    create_sample_workbook()
    excel_to_csv("multi_sheet_data.xlsx")