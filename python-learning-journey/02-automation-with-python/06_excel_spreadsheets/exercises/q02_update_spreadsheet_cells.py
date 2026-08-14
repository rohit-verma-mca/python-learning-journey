"""
Question:
Create a small sample spreadsheet with a few rows of product names and
prices. Write a program that opens it, applies a 10% price increase to
every price cell, and saves the updated file.

"""

import openpyxl

FILENAME = "products.xlsx"


def create_sample_spreadsheet():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Product", "Price"])
    sheet.append(["Notebook", 50])
    sheet.append(["Pen", 10])
    sheet.append(["Backpack", 800])
    sheet.append(["Water Bottle", 150])
    wb.save(FILENAME)
    print(f"Created sample file '{FILENAME}'.")


def increase_prices(filename, percent=10):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):  # skip header row
        price_cell = sheet.cell(row=row, column=2)
        old_price = price_cell.value
        new_price = round(old_price * (1 + percent / 100), 2)
        price_cell.value = new_price
        print(f"Row {row}: {old_price} -> {new_price}")

    wb.save(filename)
    print(f"\nUpdated prices saved to '{filename}'.")


if __name__ == "__main__":
    create_sample_spreadsheet()
    increase_prices(FILENAME, percent=10)