"""
Question:
Ask the user for a number N, then create an Excel spreadsheet containing
an N x N multiplication table, with row and column headers.

"""

import openpyxl


def create_multiplication_table(n, filename="multiplication_table.xlsx"):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Multiplication Table"

    # column headers (row 1, starting from column 2)
    for col in range(1, n + 1):
        sheet.cell(row=1, column=col + 1).value = col

    # row headers (column 1, starting from row 2) + fill in the table
    for row in range(1, n + 1):
        sheet.cell(row=row + 1, column=1).value = row
        for col in range(1, n + 1):
            sheet.cell(row=row + 1, column=col + 1).value = row * col

    wb.save(filename)
    print(f"Saved {n}x{n} multiplication table to '{filename}'.")


if __name__ == "__main__":
    n = int(input("Enter a number for the table size: "))
    create_multiplication_table(n)